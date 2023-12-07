import openpyxl


class ClsType:
    def __init__(self, name):
        self.name = name
        self.allMoney = 0

        # {'US': 0.42332031286763167, 'RU': 0.13378996424343773, 'DE': 0.11884706405818267}
        self.allCountry = {}

    def addApp(self, country, count):
        self.allMoney += count
        if country not in self.allCountry:
            self.allCountry[country] = 0
        self.allCountry[country] += count

    def print(self):
        print(self.name)
        print(self.allMoney)
        print(self.allCountry)


class ClsApp:
    def __init__(self, name, app_type):
        self.name = name
        self.app_type = app_type
        self.allMoney = 0
        self.allCountry = {}
        self.missCountry = {}

    def addApp(self, country, count):
        if country not in self.allCountry:
            self.allCountry[country] = 0
        self.allCountry[country] += count
        self.allMoney += count

    def print(self):
        print(self.allCountry)

    def printMiss(self):
        if len(self.missCountry.keys()) > 0:
            # json.dump(self.missCountry)
            print(self.name + ": " + str(self.missCountry))
            # print(self.missCountry)


wb = openpyxl.load_workbook('kakaka.xlsx')
# 获取所有工作表的名称
print(wb.sheetnames)
table = wb['Sheet2']
# 获取表格行数
nrows = table.max_row
print("表格一共有", nrows, "行")
rowCount = 498

dictTpye = {}
dictApp = {}

for row in table.iter_rows(min_row=2, max_row=rowCount):
    typeName = row[2].value
    if typeName not in dictTpye:
        dictTpye[typeName] = ClsType(typeName)
    dictTpye[typeName].addApp(row[0].value, row[3].value)

    appName = row[1].value
    if appName not in dictApp:
        dictApp[appName] = ClsApp(appName, row[2].value)
    dictApp[appName].addApp(row[0].value, row[3].value)

for app in dictTpye.values():
    for contName in list(app.allCountry.keys()):
        app.allCountry[contName] = app.allCountry[contName] / app.allMoney
        if app.allCountry[contName] < 0.1:
            del app.allCountry[contName]

# for type in dictApp.values():
#     for contName in list(type.allCountry.keys()):
#         type.allCountry[contName] = type.allCountry[contName] / type.allMoney


for app in dictApp.values():
    for appCont in dictTpye[app.app_type].allCountry.keys():
        if appCont not in app.allCountry.keys():
            app.missCountry[appCont] = dictTpye[app.app_type].allCountry[appCont]

# for type in dictApp.values():
#     type.printMiss()
wb.close()

from openpyxl import Workbook

# 创建一个工作簿对象
wb = Workbook()
# 在索引为0的位置创建一个名为mytest的sheet页
ws = wb.create_sheet('mytest', 0)
# 对sheet页设置一个颜色（16位的RGB颜色）
ws.sheet_properties.tabColor = 'ff72BA'

rowCount = 1
for app in dictApp.values():
    if len(app.missCountry.values()) > 0:
        ws.cell(row=rowCount, column=1).value = app.name
        ws.cell(row=rowCount, column=2).value = str(app.missCountry).replace("'", "").strip("{").strip("}")
        rowCount += 1
    # type.printMiss()

# 将创建的工作簿保存为Mytest.xlsx
wb.save('Mytest.xlsx')
# 最后关闭文件
wb.close()
